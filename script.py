import datetime
import sys
import time

import openpyxl


def deleteColumns(ws):
    col_to_delete = [x for x in range(1,27) if x not in [3,5,7,16]]
    cols_deleted = 0
    for i in col_to_delete:
        # print(i)
        ws.delete_cols(i - cols_deleted)
        cols_deleted = cols_deleted + 1

def deleteRows(ws, rows):
	rows_deleted = 0
	for r in rows:
		ws.delete_rows(r - rows_deleted, 1)
		rows_deleted = rows_deleted + 1

def filterRowsByStatus(ws):
	status_column = 3
	rowsToDelete = []
	i = 1

	for row in ws.iter_rows(min_row=1, min_col=status_column, max_col=status_column):
		for cell in row:
			if(cell.value != 'issued' and i != 1):
				rowsToDelete.append(i)
		i = i + 1
	return rowsToDelete
		
def filterRowsByDate(ws):
		date_column = 2
		rowsToDelete = []
		i = 1
		today = datetime.date.today()
		thirtyDateFromNow = today + datetime.timedelta(days=30)
		for row in ws.iter_rows(min_row=1, min_col=date_column, max_col=date_column):
			for cell in row:
				if(i == 1):
					continue
				rowDate = cell.value.date()
				# print(rowDate, today, thirtyDateFromNow, rowDate < today or rowDate > thirtyDateFromNow) 
				if(rowDate < today or rowDate > thirtyDateFromNow):
					rowsToDelete.append(i)
			i = i + 1
		return rowsToDelete
def formatOutputFileName():
	today = datetime.datetime.today()
	return 'ssl_' + today.strftime('%Y%m%d%H%M%S') + '.xlsx' 

def main():
	wb = openpyxl.load_workbook(sys.argv[1])
	ws = wb.active

	deleteColumns(ws)

	statusRows = filterRowsByStatus(ws)
	deleteRows(ws, statusRows)

	dateRows = filterRowsByDate(ws)
	deleteRows(ws, dateRows)

	wb.save(formatOutputFileName())

# def test():
# 	date_str = '3/25/2019'
# 	st = time.strptime(date_str, '%m/%d/%Y')
# 	dt = datetime.date(*st[:3])
# 	now = datetime.date.today()
# 	# nowand30 = now + datetime.timedelta(days=30)
# 	print(dt)

# test()
main()
